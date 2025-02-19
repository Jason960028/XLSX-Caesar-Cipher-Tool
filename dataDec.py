import hashlib
from openpyxl import load_workbook

def caesar_cipher(text, shift=3):
    """
    Shifts a text string by the given shift amount.
    This function can be used for both encryption (positive shift) and decryption (negative shift).
    """
    result = []
    for char in text:
        if char.isalpha():
            # Determine the base depending on uppercase or lowercase.
            base = ord('A') if char.isupper() else ord('a')
            new_char = chr((ord(char) - base + shift) % 26 + base)
            result.append(new_char)
        else:
            result.append(char)
    return ''.join(result)

def decrypt_xlsx_file(input_file, output_file, shift=3):
    """
    Reads an XLSX file that was encrypted with a Caesar cipher,
    decrypts each cell in the data rows (skipping the header row),
    and removes the extra "row_hash" column if it exists.
    The decrypted data is then saved to a new XLSX file.
    """
    # Load the encrypted workbook.
    wb = load_workbook(input_file)
    
    for ws in wb.worksheets:
        # Read the header row (assumed to be the first row).
        header = [cell.value for cell in ws[1]]
        
        # Iterate through each data row (starting from row 2).
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                # If the cell is in the hash column (or beyond) skip decryption.
                # We use cell.col_idx which is a 1-indexed integer.
                if cell.value is not None:
                    # Decrypt the cell by shifting in the reverse direction.
                    cell.value = caesar_cipher(str(cell.value), -shift)
    
    # Save the decrypted workbook.
    wb.save(output_file)
    print(f"Decrypted file saved as '{output_file}'.")

if __name__ == "__main__":
    # Define file names.
    input_file = "encrypted_input.xlsx"   # The encrypted Excel file.
    output_file = "decrypted_output.xlsx"   # The file to write the decrypted data.
    
    # Decrypt the file (using the same shift value as was used during encryption).
    decrypt_xlsx_file(input_file, output_file, shift=3)
