import hashlib
from openpyxl import load_workbook

def caesar_cipher(text, shift=3):
    """
    Encrypts a text string using a Caesar cipher.
    Only alphabetic characters are shifted; others remain unchanged.
    """
    result = []
    for char in text:
        if char.isalpha():
            # Determine base based on uppercase or lowercase.
            base = ord('A') if char.isupper() else ord('a')
            new_char = chr((ord(char) - base + shift) % 26 + base)
            result.append(new_char)
        else:
            result.append(char)
    return ''.join(result)


def encrypt_xlsx_file(input_file, output_file, shift=3):
    """
    Reads an XLSX file, encrypts each cell in the data rows using a Caesar cipher.
    """
    # Load the workbook.
    wb = load_workbook(input_file)
    
    for ws in wb.worksheets:
        original_max_column = ws.max_column
        
        # Process data rows (starting from row 2).
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=original_max_column):
            encrypted_values = []
            for cell in row:
                if cell.value is not None:
                    # Encrypt the cell value and update the cell.
                    encrypted_value = caesar_cipher(str(cell.value), shift)
                    cell.value = encrypted_value
                    encrypted_values.append(encrypted_value)
                else:
                    encrypted_values.append('')
    
    # Save the encrypted workbook to a new file.
    wb.save(output_file)
    print(f"Encrypted file saved as '{output_file}'.")


if __name__ == "__main__":
    """
    Change input_file name to the file you want to encrypt.
    """
    input_file = "sample.xlsx"            
    output_file = "encrypted_sample.xlsx" 
    
    encrypt_xlsx_file(input_file, output_file, shift=3)
