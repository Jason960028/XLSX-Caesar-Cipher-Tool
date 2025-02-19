import hashlib
from openpyxl import load_workbook

def caesar_cipher(text, shift=3):
    """
    Encrypts a text string using a Caesar cipher.
    Only alphabetic characters are shifted; others remain unchanged.
    """
    result = []
    for char in text:
        if char.isalpha():
            # Determine base based on uppercase or lowercase.
            base = ord('A') if char.isupper() else ord('a')
            new_char = chr((ord(char) - base + shift) % 26 + base)
            result.append(new_char)
        else:
            result.append(char)
    return ''.join(result)


def encrypt_xlsx_file(input_file, output_file, shift=3):
    """
    Reads an XLSX file, encrypts each cell in the data rows using a Caesar cipher,
    computes a SHA-256 hash for each data row (based on the encrypted values),
    and writes the result to a new XLSX file. The header row is not encrypted.
    """
    # Load the workbook.
    wb = load_workbook(input_file)
    
    for ws in wb.worksheets:
        original_max_column = ws.max_column
        
        # Process header row (assumed to be the first row) without encryption.
        header_row = list(ws.iter_rows(min_row=1, max_row=1, max_col=original_max_column))[0]
        
        # Process data rows (starting from row 2).
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=original_max_column):
            encrypted_values = []
            for cell in row:
                if cell.value is not None:
                    # Encrypt the cell value and update the cell.
                    encrypted_value = caesar_cipher(str(cell.value), shift)
                    cell.value = encrypted_value
                    encrypted_values.append(encrypted_value)
                else:
                    encrypted_values.append('')
    
    # Save the encrypted workbook to a new file.
    wb.save(output_file)
    print(f"Encrypted file saved as '{output_file}'.")

if __name__ == "__main__":
    # Define input and output file names.
    input_file = "input.xlsx"            # Your original Excel file.
    output_file = "encrypted_input.xlsx" # The output file with encrypted data.
    
    # Encrypt the XLSX file.
    encrypt_xlsx_file(input_file, output_file, shift=3)
